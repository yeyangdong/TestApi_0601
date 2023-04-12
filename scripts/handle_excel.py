# Excel 第四节课,封装调用Excel

from openpyxl import load_workbook


class HandleExcel:

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        wb = load_workbook(self.filename)
        ws = wb[self.sheetname]

        testcase_list = []
        headers_list = []
        for row in range(1, ws.max_row + 1):
            # 存放每一行的用例数据
            one_row_dict = {}
            for column in range(1, ws.max_column + 1):
                one_cell_value = ws.cell(row, column).value
                if row == 1:
                    headers_list.append(one_cell_value)
                else:
                    key = headers_list[column - 1]  # 将上面的headers_list拿过来做字典中的key
                    one_row_dict[key] = one_cell_value  # one_row_dict的key,赋值为one_cell_value的值
                    testcase_list.append(one_row_dict)

        return testcase_list

    def write_data(self, row, column, data):
        # 指定在某一行某一列写数据
        # 将数据写入到Excel中,不能与读取时共用一个workbook对象
        wb = load_workbook(self.filename)
        ws = wb[self.sheetname]

        # 第一种写入方式
        # one_cell = ws.cell(row, column)
        # one_cell.value = data

        # 第二种写入方式
        ws.cell(row, column, value=data)

        wb.save(self.filename)


if __name__ == '__main__':
    excel_filename = "yaml/py05.xlsx"
    sheet_name = "登陆1"
    do_excel = HandleExcel(excel_filename, sheet_name)
    res = do_excel.read_data()

    # print(do_excel.read_data())#,读数据

    do_excel.write_data(2, 6, 1)  # 写数据
